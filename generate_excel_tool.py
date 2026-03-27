"""
Generate the Samples Management Excel Tool (.xlsx)
100% formula-based, no VBA/macros required.
Compatible with Excel 365, Excel 2021+, and Excel Online.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection, numbers
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from copy import copy


def add_named_range(wb, name, attr_text):
    """Helper to add a named range compatible with openpyxl 3.x"""
    dn = DefinedName(name, attr_text=attr_text)
    wb.defined_names.add(dn)

# ============================================================
# CONSTANTS (mirroring app.js)
# ============================================================

PRODUCT_TYPES = ["T-Shirt", "Sweatshirt", "Trousers"]
PRODUCT_COLOURS = ["White", "Orange", "Yellow", "Blue", "Green"]
SIZES_ALPHA = ["XXS", "XS", "S", "M", "L", "XL", "XXL"]
SIZES_NUMERIC = ["36", "38", "40", "42", "44", "46", "48", "50"]
TARGET_MARKETS = ["North", "South"]

COLOUR_FILLS = {
    "White":  PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    "Orange": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "Yellow": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "Blue":   PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
    "Green":  PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
}

# ============================================================
# STYLES
# ============================================================

FONT_TITLE = Font(name="Aptos", size=16, bold=True, color="333333")
FONT_SECTION = Font(name="Aptos", size=12, bold=True, color="333333")
FONT_LABEL = Font(name="Aptos", size=11, bold=True, color="555555")
FONT_NORMAL = Font(name="Aptos", size=11, color="333333")
FONT_INPUT = Font(name="Aptos", size=11, color="000000")
FONT_SMALL = Font(name="Aptos", size=9, color="999999")
FONT_TOTAL = Font(name="Aptos", size=14, bold=True, color="1565C0")
FONT_HEADER = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
FONT_FORMULA = Font(name="Aptos", size=10, color="888888")

FILL_INPUT = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
FILL_HEADER = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
FILL_SECTION = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
FILL_ZERO = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FILL_TOTAL_ROW = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
FILL_DESTRUCTIVE = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
FILL_SAFE = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_REF_HEADER = PatternFill(start_color="546E7A", end_color="546E7A", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
BORDER_INPUT = Border(
    left=Side(style="thin", color="1565C0"),
    right=Side(style="thin", color="1565C0"),
    top=Side(style="thin", color="1565C0"),
    bottom=Side(style="medium", color="1565C0"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

PROTECTION_LOCKED = Protection(locked=True)
PROTECTION_UNLOCKED = Protection(locked=False)


def apply_style(cell, font=None, fill=None, border=None, alignment=None, protection=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if protection:
        cell.protection = protection
    if number_format:
        cell.number_format = number_format


# ============================================================
# SHEET: REF (Reference Data)
# ============================================================

def create_ref_sheet(wb):
    ws = wb.create_sheet("REF")
    ws.sheet_properties.tabColor = "78909C"

    # --- Column A: Product Types ---
    ws["A1"] = "Product Types"
    apply_style(ws["A1"], font=FONT_HEADER, fill=FILL_REF_HEADER, alignment=ALIGN_CENTER)
    for i, pt in enumerate(PRODUCT_TYPES, start=2):
        ws.cell(row=i, column=1, value=pt)
        apply_style(ws.cell(row=i, column=1), font=FONT_NORMAL)

    # --- Column C: Product Colours ---
    ws["C1"] = "Colours"
    apply_style(ws["C1"], font=FONT_HEADER, fill=FILL_REF_HEADER, alignment=ALIGN_CENTER)
    for i, c in enumerate(PRODUCT_COLOURS, start=2):
        ws.cell(row=i, column=3, value=c)
        apply_style(ws.cell(row=i, column=3), font=FONT_NORMAL)
        if c in COLOUR_FILLS:
            ws.cell(row=i, column=3).fill = COLOUR_FILLS[c]

    # --- Column E: Sizes Alpha ---
    ws["E1"] = "Sizes Alpha"
    apply_style(ws["E1"], font=FONT_HEADER, fill=FILL_REF_HEADER, alignment=ALIGN_CENTER)
    for i, s in enumerate(SIZES_ALPHA, start=2):
        ws.cell(row=i, column=5, value=s)
        apply_style(ws.cell(row=i, column=5), font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # --- Column G: Sizes Numeric ---
    ws["G1"] = "Sizes Numeric"
    apply_style(ws["G1"], font=FONT_HEADER, fill=FILL_REF_HEADER, alignment=ALIGN_CENTER)
    for i, s in enumerate(SIZES_NUMERIC, start=2):
        ws.cell(row=i, column=7, value=s)
        apply_style(ws.cell(row=i, column=7), font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # --- Column I: Target Markets ---
    ws["I1"] = "Markets"
    apply_style(ws["I1"], font=FONT_HEADER, fill=FILL_REF_HEADER, alignment=ALIGN_CENTER)
    for i, m in enumerate(TARGET_MARKETS, start=2):
        ws.cell(row=i, column=9, value=m)
        apply_style(ws.cell(row=i, column=9), font=FONT_NORMAL)

    # --- Column K: YES/NO ---
    ws["K1"] = "YesNo"
    apply_style(ws["K1"], font=FONT_HEADER, fill=FILL_REF_HEADER, alignment=ALIGN_CENTER)
    ws.cell(row=2, column=11, value="YES")
    ws.cell(row=3, column=11, value="NO")

    # --- Column M: Size System Names ---
    ws["M1"] = "Size Systems"
    apply_style(ws["M1"], font=FONT_HEADER, fill=FILL_REF_HEADER, alignment=ALIGN_CENTER)
    ws.cell(row=2, column=13, value="Alpha (XXS - XXL)")
    ws.cell(row=3, column=13, value="Numeric (36 - 50)")

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["M"].width = 22

    # Define named ranges
    add_named_range(wb,"ProductTypes", attr_text="REF!$A$2:$A$4")
    add_named_range(wb,"Colours", attr_text="REF!$C$2:$C$6")
    add_named_range(wb,"SizesAlpha", attr_text="REF!$E$2:$E$8")
    add_named_range(wb,"SizesNumeric", attr_text="REF!$G$2:$G$9")
    add_named_range(wb,"Markets", attr_text="REF!$I$2:$I$3")
    add_named_range(wb,"YesNo", attr_text="REF!$K$2:$K$3")
    add_named_range(wb,"SizeSystems", attr_text="REF!$M$2:$M$3")

    return ws


# ============================================================
# SHEET: CONFIG (User Configuration)
# ============================================================

def create_config_sheet(wb):
    ws = wb.create_sheet("CONFIG")
    ws.sheet_properties.tabColor = "1565C0"

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 22

    # ---- TITLE ----
    ws.merge_cells("B2:C2")
    ws["B2"] = "SAMPLES MANAGEMENT"
    apply_style(ws["B2"], font=FONT_TITLE)

    ws.merge_cells("B3:C3")
    ws["B3"] = "Configuration Sheet - Fill in the yellow cells"
    apply_style(ws["B3"], font=FONT_SMALL)

    # ---- SECTION: PROJECT INFO ----
    row = 5
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = "PROJECT INFORMATION"
    apply_style(ws[f"B{row}"], font=FONT_SECTION, fill=FILL_SECTION)
    apply_style(ws[f"C{row}"], fill=FILL_SECTION)

    row = 6
    ws[f"B{row}"] = "Collection Name"
    apply_style(ws[f"B{row}"], font=FONT_LABEL)
    ws[f"C{row}"] = ""
    apply_style(ws[f"C{row}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                protection=PROTECTION_UNLOCKED)

    row = 7
    ws[f"B{row}"] = "Product Type"
    apply_style(ws[f"B{row}"], font=FONT_LABEL)
    ws[f"C{row}"] = ""
    apply_style(ws[f"C{row}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                protection=PROTECTION_UNLOCKED)
    dv_product = DataValidation(type="list", formula1="ProductTypes", allow_blank=True)
    dv_product.prompt = "Select a product type"
    dv_product.promptTitle = "Product Type"
    ws.add_data_validation(dv_product)
    dv_product.add(ws[f"C{row}"])

    # ---- SECTION: COLOURS ----
    row = 9
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = "PRODUCT COLOURS"
    apply_style(ws[f"B{row}"], font=FONT_SECTION, fill=FILL_SECTION)
    apply_style(ws[f"C{row}"], fill=FILL_SECTION)

    ws[f"B10"] = "Select YES for each colour"
    apply_style(ws[f"B10"], font=FONT_SMALL)

    dv_yesno = DataValidation(type="list", formula1="YesNo", allow_blank=True)
    ws.add_data_validation(dv_yesno)

    colour_rows = {}
    for i, colour in enumerate(PRODUCT_COLOURS):
        r = 11 + i
        colour_rows[colour] = r
        ws[f"B{r}"] = colour
        apply_style(ws[f"B{r}"], font=FONT_LABEL)
        if colour in COLOUR_FILLS:
            ws[f"B{r}"].fill = COLOUR_FILLS[colour]
        ws[f"C{r}"] = "NO"
        apply_style(ws[f"C{r}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                    alignment=ALIGN_CENTER, protection=PROTECTION_UNLOCKED)
        dv_yesno.add(ws[f"C{r}"])

    # Colour count (named cell for formulas)
    # C11:C15 = colour YES/NO cells
    row = 17
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = "SHOOTING & SIZES"
    apply_style(ws[f"B{row}"], font=FONT_SECTION, fill=FILL_SECTION)
    apply_style(ws[f"C{row}"], fill=FILL_SECTION)

    row = 18
    ws[f"B{row}"] = "Colour for Shooting"
    apply_style(ws[f"B{row}"], font=FONT_LABEL)
    ws[f"C{row}"] = ""
    apply_style(ws[f"C{row}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                protection=PROTECTION_UNLOCKED)
    # Dynamic validation: only colours with YES
    # We use OFFSET+MATCH trick or just list the 5 colours (user must pick one they selected)
    dv_shooting = DataValidation(type="list", formula1="Colours", allow_blank=True)
    dv_shooting.prompt = "Select a colour from those selected above"
    dv_shooting.promptTitle = "Shooting Colour"
    ws.add_data_validation(dv_shooting)
    dv_shooting.add(ws[f"C{row}"])

    row = 19
    ws[f"B{row}"] = "Size System"
    apply_style(ws[f"B{row}"], font=FONT_LABEL)
    ws[f"C{row}"] = ""
    apply_style(ws[f"C{row}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                protection=PROTECTION_UNLOCKED)
    dv_sizes = DataValidation(type="list", formula1="SizeSystems", allow_blank=True)
    dv_sizes.prompt = "Select Alpha or Numeric"
    dv_sizes.promptTitle = "Size System"
    ws.add_data_validation(dv_sizes)
    dv_sizes.add(ws[f"C{row}"])

    row = 20
    ws[f"B{row}"] = "Base Size"
    apply_style(ws[f"B{row}"], font=FONT_LABEL)
    ws[f"C{row}"] = ""
    apply_style(ws[f"C{row}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                protection=PROTECTION_UNLOCKED)
    # Dynamic: depends on size system chosen
    # Use INDIRECT to switch between SizesAlpha and SizesNumeric
    dv_base = DataValidation(
        type="list",
        formula1='=IF(C19="Alpha (XXS - XXL)",SizesAlpha,SizesNumeric)',
        allow_blank=True
    )
    dv_base.prompt = "Select the base size (depends on size system)"
    dv_base.promptTitle = "Base Size"
    ws.add_data_validation(dv_base)
    dv_base.add(ws[f"C{row}"])

    # ---- SECTION: TARGET MARKETS ----
    row = 22
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = "TARGET MARKETS"
    apply_style(ws[f"B{row}"], font=FONT_SECTION, fill=FILL_SECTION)
    apply_style(ws[f"C{row}"], fill=FILL_SECTION)

    dv_yesno2 = DataValidation(type="list", formula1="YesNo", allow_blank=True)
    ws.add_data_validation(dv_yesno2)

    market_rows = {}
    for i, market in enumerate(TARGET_MARKETS):
        r = 23 + i
        market_rows[market] = r
        ws[f"B{r}"] = market
        apply_style(ws[f"B{r}"], font=FONT_LABEL)
        ws[f"C{r}"] = "NO"
        apply_style(ws[f"C{r}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                    alignment=ALIGN_CENTER, protection=PROTECTION_UNLOCKED)
        dv_yesno2.add(ws[f"C{r}"])

    # ---- SECTION: SPECIFIC TESTS ----
    row = 26
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = "SPECIFIC TESTS"
    apply_style(ws[f"B{row}"], font=FONT_SECTION, fill=FILL_SECTION)
    apply_style(ws[f"C{row}"], fill=FILL_SECTION)

    dv_yesno3 = DataValidation(type="list", formula1="YesNo", allow_blank=True)
    ws.add_data_validation(dv_yesno3)

    row = 27
    ws[f"B{row}"] = "High visibility"
    apply_style(ws[f"B{row}"], font=FONT_LABEL)
    ws[f"C{row}"] = "NO"
    apply_style(ws[f"C{row}"], font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_INPUT,
                alignment=ALIGN_CENTER, protection=PROTECTION_UNLOCKED)
    dv_yesno3.add(ws[f"C{row}"])

    # ---- RIGHT COLUMN: CALCULATED SUMMARY ----
    row = 5
    ws.merge_cells(f"E{row}:F{row}")
    ws[f"E{row}"] = "CALCULATED SUMMARY"
    apply_style(ws[f"E{row}"], font=FONT_SECTION, fill=FILL_SECTION)
    apply_style(ws[f"F{row}"], fill=FILL_SECTION)

    # Nb Colours (C)
    row = 7
    ws[f"E{row}"] = "Nb colours selected (C)"
    apply_style(ws[f"E{row}"], font=FONT_LABEL)
    ws[f"F{row}"] = '=COUNTIF(C11:C15,"YES")'
    apply_style(ws[f"F{row}"], font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Nb Markets (M)
    row = 8
    ws[f"E{row}"] = "Nb markets selected (M)"
    apply_style(ws[f"E{row}"], font=FONT_LABEL)
    ws[f"F{row}"] = '=COUNTIF(C23:C24,"YES")'
    apply_style(ws[f"F{row}"], font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Specific tests count
    row = 9
    ws[f"E{row}"] = "Specific tests (tSpec)"
    apply_style(ws[f"E{row}"], font=FONT_LABEL)
    ws[f"F{row}"] = '=IF(C27="YES",1,0)'
    apply_style(ws[f"F{row}"], font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Formula display
    row = 11
    ws[f"E{row}"] = "Formula: C×M + M + 1 + tSpec"
    apply_style(ws[f"E{row}"], font=FONT_FORMULA)

    # Wash tests
    row = 13
    ws[f"E{row}"] = "Wash tests (C × M)"
    apply_style(ws[f"E{row}"], font=FONT_LABEL)
    ws[f"F{row}"] = "=F7*F8"
    apply_style(ws[f"F{row}"], font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Personalization guides
    row = 14
    ws[f"E{row}"] = "Personalization guides (M)"
    apply_style(ws[f"E{row}"], font=FONT_LABEL)
    ws[f"F{row}"] = "=F8"
    apply_style(ws[f"F{row}"], font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Photo shooting
    row = 15
    ws[f"E{row}"] = "Photo shooting"
    apply_style(ws[f"E{row}"], font=FONT_LABEL)
    ws[f"F{row}"] = 1
    apply_style(ws[f"F{row}"], font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Specific tests
    row = 16
    ws[f"E{row}"] = "Specific tests"
    apply_style(ws[f"E{row}"], font=FONT_LABEL)
    ws[f"F{row}"] = "=F9"
    apply_style(ws[f"F{row}"], font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # TOTAL
    row = 18
    ws.merge_cells(f"E{row}:E{row}")
    ws[f"E{row}"] = "TOTAL SAMPLES"
    apply_style(ws[f"E{row}"], font=Font(name="Aptos", size=14, bold=True, color="1565C0"))
    ws[f"F{row}"] = "=F13+F14+F15+F16"
    apply_style(ws[f"F{row}"], font=FONT_TOTAL, alignment=ALIGN_CENTER,
                fill=PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
                border=Border(
                    left=Side(style="medium", color="1565C0"),
                    right=Side(style="medium", color="1565C0"),
                    top=Side(style="medium", color="1565C0"),
                    bottom=Side(style="medium", color="1565C0"),
                ))

    # ---- Helper cells for SAMPLES sheet ----
    # Selected colours list (auxiliary column E, rows 21-25)
    row = 20
    ws[f"E{row}"] = "Selected colours (helper)"
    apply_style(ws[f"E{row}"], font=FONT_SMALL)

    # Build a list of selected colours using nested IFs
    # E21 = 1st selected colour, E22 = 2nd, etc.
    # We use SMALL+IF array or simple sequential approach
    for i in range(5):
        r = 21 + i
        colour_name = PRODUCT_COLOURS[i]
        ws[f"E{r}"] = f'=IF(C{11+i}="YES",B{11+i},"")'
        apply_style(ws[f"E{r}"], font=FONT_SMALL)

    # Compact list of selected colours (no gaps) using FILTER-like approach
    # For broad compat, we'll place in column E rows 27-31
    row = 26
    ws[f"E{row}"] = "Compact colour list (helper)"
    apply_style(ws[f"E{row}"], font=FONT_SMALL)

    # Using IFERROR+SMALL+IF array formula approach for compatibility
    # But for Excel 365, FILTER is simpler. Let's use FILTER.
    for i in range(5):
        r = 27 + i
        # IFERROR(INDEX(FILTER(...),...), "")
        ws[f"E{r}"] = f'=IFERROR(INDEX(FILTER(B11:B15,C11:C15="YES"),{i+1}),"")'
        apply_style(ws[f"E{r}"], font=FONT_SMALL)

    # Selected markets helper
    row = 33
    ws[f"E{row}"] = "Selected markets (helper)"
    apply_style(ws[f"E{row}"], font=FONT_SMALL)
    for i in range(2):
        r = 34 + i
        ws[f"E{r}"] = f'=IFERROR(INDEX(FILTER(B23:B24,C23:C24="YES"),{i+1}),"")'
        apply_style(ws[f"E{r}"], font=FONT_SMALL)

    # Size helpers
    row = 37
    ws[f"E{row}"] = "Size helpers"
    apply_style(ws[f"E{row}"], font=FONT_SMALL)

    # Min size
    ws[f"E38"] = "Size Min"
    ws[f"F38"] = '=IF(C19="Alpha (XXS - XXL)",REF!E2,REF!G2)'
    apply_style(ws[f"E38"], font=FONT_SMALL)
    apply_style(ws[f"F38"], font=FONT_SMALL)

    # Max size
    ws[f"E39"] = "Size Max"
    ws[f"F39"] = '=IF(C19="Alpha (XXS - XXL)",REF!E8,REF!G9)'
    apply_style(ws[f"E39"], font=FONT_SMALL)
    apply_style(ws[f"F39"], font=FONT_SMALL)

    # Base size (copy from input)
    ws[f"E40"] = "Size Base"
    ws[f"F40"] = "=C20"
    apply_style(ws[f"E40"], font=FONT_SMALL)
    apply_style(ws[f"F40"], font=FONT_SMALL)

    # All sizes for the selected system (for ORDER_GRID headers)
    row = 42
    ws[f"E{row}"] = "All sizes for system (helper)"
    apply_style(ws[f"E{row}"], font=FONT_SMALL)
    for i in range(8):
        r = 43 + i
        ws[f"E{r}"] = f'=IF(C19="Alpha (XXS - XXL)",IF({i+1}<=7,INDEX(REF!E2:E8,{i+1}),""),INDEX(REF!G2:G9,{i+1}))'
        apply_style(ws[f"E{r}"], font=FONT_SMALL)

    # Named ranges for key cells
    add_named_range(wb,"NbColours", attr_text="CONFIG!$F$7")
    add_named_range(wb,"NbMarkets", attr_text="CONFIG!$F$8")
    add_named_range(wb,"TotalSamples", attr_text="CONFIG!$F$18")
    add_named_range(wb,"SizeMin", attr_text="CONFIG!$F$38")
    add_named_range(wb,"SizeMax", attr_text="CONFIG!$F$39")
    add_named_range(wb,"SizeBase", attr_text="CONFIG!$F$40")
    add_named_range(wb,"ShootingColour", attr_text="CONFIG!$C$18")
    add_named_range(wb,"SelectedColours", attr_text="CONFIG!$E$27:$E$31")
    add_named_range(wb,"SelectedMarkets", attr_text="CONFIG!$E$34:$E$35")
    add_named_range(wb,"HighVisibility", attr_text="CONFIG!$C$27")
    add_named_range(wb,"AllSizes", attr_text="CONFIG!$E$43:$E$50")

    # Conditional formatting: warning if shooting colour not among selected
    ws.conditional_formatting.add(
        "C18",
        FormulaRule(
            formula=['AND(C18<>"",COUNTIF(E27:E31,C18)=0)'],
            fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
            font=Font(color="C62828")
        )
    )

    # Freeze panes
    ws.freeze_panes = "B5"

    return ws


# ============================================================
# SHEET: SAMPLES (Generated Sample Instances)
# ============================================================

def create_samples_sheet(wb):
    ws = wb.create_sheet("SAMPLES")
    ws.sheet_properties.tabColor = "FF8F00"

    MAX_SAMPLES = 30  # Maximum rows to support

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18

    # Headers
    headers = ["#", "Colour", "Size", "Source", "Destructive", "Market"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        apply_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

    # Sample generation formulas
    # Key references:
    # CONFIG!F7 = NbColours (C), CONFIG!F8 = NbMarkets (M)
    # CONFIG!F9 = tSpec
    # CONFIG!E27:E31 = compact selected colours
    # CONFIG!E34:E35 = compact selected markets
    # CONFIG!F38 = SizeMin, CONFIG!F39 = SizeMax, CONFIG!F40 = SizeBase
    # CONFIG!C18 = ShootingColour

    # Boundary calculations (in column H, hidden helper)
    ws["H1"] = "WashEnd"
    apply_style(ws["H1"], font=FONT_SMALL)
    ws["H2"] = "=CONFIG!$F$7*CONFIG!$F$8"  # C*M

    ws["I1"] = "PersoEnd"
    apply_style(ws["I1"], font=FONT_SMALL)
    ws["I2"] = "=H2+CONFIG!$F$8"  # C*M + M

    ws["J1"] = "ShootEnd"
    apply_style(ws["J1"], font=FONT_SMALL)
    ws["J2"] = "=I2+1"  # C*M + M + 1

    ws["K1"] = "Total"
    apply_style(ws["K1"], font=FONT_SMALL)
    ws["K2"] = "=J2+CONFIG!$F$9"  # C*M + M + 1 + tSpec

    for row in range(2, MAX_SAMPLES + 2):
        n = row - 1  # sample number (1-based)

        # Column A: Sample # (show only if within total)
        ws.cell(row=row, column=1).value = f'=IF({n}<=K$2,{n},"")'
        apply_style(ws.cell(row=row, column=1), font=FONT_NORMAL, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

        # Column B: Colour
        # Wash tests (1 to C*M): cycle through colours
        #   colour_index = MOD(n-1, C) + 1 → INDEX(SelectedColours, colour_index)
        # Perso guide (C*M+1 to C*M+M): cycle through colours
        #   colour_index = MOD(n-H$2-1, C) + 1
        # Shooting (C*M+M+1): ShootingColour
        # Specific tests: first selected colour
        colour_formula = (
            f'=IF({n}>K$2,"",'
            f'IF({n}<=H$2,'
            f'INDEX(CONFIG!$E$27:$E$31,MOD({n}-1,CONFIG!$F$7)+1),'
            f'IF({n}<=I$2,'
            f'INDEX(CONFIG!$E$27:$E$31,MOD({n}-H$2-1,CONFIG!$F$7)+1),'
            f'IF({n}<=J$2,'
            f'CONFIG!$C$18,'
            f'INDEX(CONFIG!$E$27:$E$31,1)'
            f'))))'
        )
        ws.cell(row=row, column=2).value = colour_formula
        apply_style(ws.cell(row=row, column=2), font=FONT_NORMAL, border=BORDER_THIN)

        # Column C: Size
        # Wash tests: first wash test = SizeMax (if SizeMax <> SizeBase), rest = SizeBase
        # Perso guide: SizeMin
        # Shooting: SizeBase
        # Specific tests (High vis): SizeMin
        size_formula = (
            f'=IF({n}>K$2,"",'
            f'IF({n}<=H$2,'
            f'IF(AND({n}=1,CONFIG!$F$39<>CONFIG!$F$40),CONFIG!$F$39,CONFIG!$F$40),'
            f'IF({n}<=I$2,'
            f'CONFIG!$F$38,'
            f'IF({n}<=J$2,'
            f'CONFIG!$F$40,'
            f'CONFIG!$F$38'
            f'))))'
        )
        ws.cell(row=row, column=3).value = size_formula
        apply_style(ws.cell(row=row, column=3), font=FONT_NORMAL, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

        # Column D: Source
        source_formula = (
            f'=IF({n}>K$2,"",'
            f'IF({n}<=H$2,"Wash tests",'
            f'IF({n}<=I$2,"Personalization guide",'
            f'IF({n}<=J$2,"Photo shooting",'
            f'"High visibility"'
            f'))))'
        )
        ws.cell(row=row, column=4).value = source_formula
        apply_style(ws.cell(row=row, column=4), font=FONT_NORMAL, border=BORDER_THIN)

        # Column E: Destructive
        destr_formula = (
            f'=IF({n}>K$2,"",'
            f'IF({n}<=H$2,"YES",'
            f'IF({n}<=I$2,"YES",'
            f'IF({n}<=J$2,"NO",'
            f'"YES"'
            f'))))'
        )
        ws.cell(row=row, column=5).value = destr_formula
        apply_style(ws.cell(row=row, column=5), font=FONT_NORMAL, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

        # Column F: Market (for wash tests and perso guide)
        # Wash tests: market_index = QUOTIENT(n-1, C) + 1
        # Perso guide: market_index = n - C*M
        market_formula = (
            f'=IF({n}>K$2,"",'
            f'IF({n}<=H$2,'
            f'IFERROR(INDEX(CONFIG!$E$34:$E$35,QUOTIENT({n}-1,CONFIG!$F$7)+1),""),'
            f'IF({n}<=I$2,'
            f'IFERROR(INDEX(CONFIG!$E$34:$E$35,{n}-H$2),""),'
            f'""'
            f')))'
        )
        ws.cell(row=row, column=6).value = market_formula
        apply_style(ws.cell(row=row, column=6), font=FONT_NORMAL, border=BORDER_THIN)

    # Conditional formatting: destructive YES = red, NO = green
    ws.conditional_formatting.add(
        f"E2:E{MAX_SAMPLES+1}",
        CellIsRule(operator="equal", formula=['"YES"'],
                   fill=FILL_DESTRUCTIVE, font=Font(color="C62828"))
    )
    ws.conditional_formatting.add(
        f"E2:E{MAX_SAMPLES+1}",
        CellIsRule(operator="equal", formula=['"NO"'],
                   fill=FILL_SAFE, font=Font(color="2E7D32"))
    )

    # Conditional formatting: colour cells get background based on colour name
    for colour, fill in COLOUR_FILLS.items():
        ws.conditional_formatting.add(
            f"B2:B{MAX_SAMPLES+1}",
            CellIsRule(operator="equal", formula=[f'"{colour}"'], fill=fill)
        )

    # Conditional formatting: hide rows beyond total (light gray if empty)
    ws.conditional_formatting.add(
        f"A2:F{MAX_SAMPLES+1}",
        FormulaRule(
            formula=[f'$A2=""'],
            font=Font(color="FFFFFF"),
            fill=FILL_WHITE
        )
    )

    return ws


# ============================================================
# SHEET: ORDER_GRID (Colour × Size Matrix)
# ============================================================

def create_order_grid_sheet(wb):
    ws = wb.create_sheet("ORDER_GRID")
    ws.sheet_properties.tabColor = "2E7D32"

    ws.column_dimensions["A"].width = 16

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "SAMPLES TO ORDER"
    apply_style(ws["A1"], font=FONT_TITLE)

    ws.merge_cells("A2:I2")
    ws["A2"] = "Colour × Size matrix - automatically calculated from configuration"
    apply_style(ws["A2"], font=FONT_SMALL)

    # Headers row 4: sizes from selected system
    # Column A = colour label, Columns B-I = sizes, last col = Total
    ws["A4"] = ""
    apply_style(ws["A4"], font=FONT_HEADER, fill=FILL_HEADER, border=BORDER_THIN)

    for i in range(8):  # max 8 sizes
        col = i + 2  # B=2, C=3, ...
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 10
        # Header = size name from CONFIG helpers
        cell = ws.cell(row=4, column=col)
        cell.value = f'=IF(CONFIG!E{43+i}="","",CONFIG!E{43+i})'
        apply_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

    # Total column (J)
    ws.column_dimensions["J"].width = 10
    ws["J4"] = "Total"
    apply_style(ws["J4"], font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER,
                border=BORDER_THIN)

    # Data rows (5 colours max, rows 5-9)
    for c_idx in range(5):
        row = 5 + c_idx
        # Column A: colour name
        ws.cell(row=row, column=1).value = f'=IF(CONFIG!E{27+c_idx}="","",CONFIG!E{27+c_idx})'
        apply_style(ws.cell(row=row, column=1), font=Font(name="Aptos", size=11, bold=True),
                    border=BORDER_THIN)

        # Columns B-I: COUNTIFS on SAMPLES sheet
        for s_idx in range(8):
            col = s_idx + 2
            cell = ws.cell(row=row, column=col)
            # Count samples where colour matches AND size matches
            # Only count if both header and row label are non-empty
            size_ref = f"CONFIG!$E${43+s_idx}"
            colour_ref = f"$A{row}"
            cell.value = (
                f'=IF(OR({colour_ref}="",{size_ref}=""),"",'
                f'COUNTIFS(SAMPLES!$B$2:$B$31,{colour_ref},SAMPLES!$C$2:$C$31,{size_ref}))'
            )
            apply_style(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)

        # Column J: row total
        ws.cell(row=row, column=10).value = (
            f'=IF($A{row}="","",SUMPRODUCT((SAMPLES!$B$2:$B$31=$A{row})*(SAMPLES!$A$2:$A$31<>"")))'
        )
        apply_style(ws.cell(row=row, column=10), font=Font(name="Aptos", size=11, bold=True),
                    alignment=ALIGN_CENTER, border=BORDER_THIN,
                    fill=FILL_TOTAL_ROW)

    # Total row (row 10)
    row = 10
    ws.cell(row=row, column=1, value="Total")
    apply_style(ws.cell(row=row, column=1), font=Font(name="Aptos", size=11, bold=True),
                fill=FILL_TOTAL_ROW, border=BORDER_THIN)

    for s_idx in range(8):
        col = s_idx + 2
        col_letter = get_column_letter(col)
        cell = ws.cell(row=row, column=col)
        cell.value = (
            f'=IF({col_letter}4="","",'
            f'SUMPRODUCT((SAMPLES!$C$2:$C$31=CONFIG!$E${43+s_idx})*(SAMPLES!$A$2:$A$31<>"")))'
        )
        apply_style(cell, font=Font(name="Aptos", size=11, bold=True),
                    alignment=ALIGN_CENTER, fill=FILL_TOTAL_ROW, border=BORDER_THIN)

    # Grand total
    ws.cell(row=row, column=10).value = "=CONFIG!F18"
    apply_style(ws.cell(row=row, column=10),
                font=Font(name="Aptos", size=12, bold=True, color="1565C0"),
                alignment=ALIGN_CENTER,
                fill=PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
                border=Border(
                    left=Side(style="medium", color="1565C0"),
                    right=Side(style="medium", color="1565C0"),
                    top=Side(style="medium", color="1565C0"),
                    bottom=Side(style="medium", color="1565C0"),
                ))

    # Conditional formatting: zero cells in gray
    ws.conditional_formatting.add(
        "B5:I9",
        CellIsRule(operator="equal", formula=["0"],
                   fill=FILL_ZERO, font=Font(color="CCCCCC"))
    )

    # Conditional formatting: hide empty cells
    ws.conditional_formatting.add(
        "A5:J9",
        FormulaRule(
            formula=['$A5=""'],
            font=Font(color="FFFFFF"),
            fill=FILL_WHITE,
            border=Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="thin", color="FFFFFF"),
            )
        )
    )
    ws.conditional_formatting.add(
        "B4:I4",
        FormulaRule(
            formula=['B4=""'],
            font=Font(color="37474F"),
            fill=FILL_HEADER
        )
    )

    return ws


# ============================================================
# SHEET: SAMPLE_FLOW (Action Flow with Stock Tracking)
# ============================================================

def create_sample_flow_sheet(wb):
    ws = wb.create_sheet("SAMPLE_FLOW")
    ws.sheet_properties.tabColor = "C62828"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 16

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "SAMPLE FLOW"
    apply_style(ws["A1"], font=FONT_TITLE)

    ws.merge_cells("A2:E2")
    ws["A2"] = "Sequential tracking of sample usage and remaining stock"
    apply_style(ws["A2"], font=FONT_SMALL)

    # Headers row 4
    headers = ["#", "Action", "Destructive", "Samples Used (Colour / Size)", "Remaining Stock"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        apply_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

    # The action flow follows this order:
    # 1. Size set control (non-destructive) - uses samples at SizeMin, SizeBase, SizeMax
    # 2. Wash tests (destructive) - all wash test samples
    # 3. Personalization guide (destructive) - all perso samples
    # 4. Specific tests if any (destructive for High vis)
    # 5. Photo shooting (non-destructive) - last

    # We'll build this with static action rows and formulas
    # Row 5: Size set control
    # Row 6: Wash tests
    # Row 7: Personalization guide
    # Row 8: High visibility (conditional)
    # Row 9: Photo shooting

    actions = [
        ("Size set control", "NO"),
        ("Wash tests", "YES"),
        ("Personalization guide", "YES"),
        ("High visibility", "YES"),
        ("Photo shooting", "NO"),
    ]

    for i, (action_name, destructive) in enumerate(actions):
        row = 5 + i
        # Column A: #
        if action_name == "High visibility":
            ws.cell(row=row, column=1).value = f'=IF(CONFIG!C27="YES",{i+1},"")'
        elif action_name == "Photo shooting":
            ws.cell(row=row, column=1).value = f'=IF(CONFIG!C27="YES",{i+1},{i})'
        else:
            ws.cell(row=row, column=1).value = i + 1
        apply_style(ws.cell(row=row, column=1), font=FONT_NORMAL, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

        # Column B: Action name
        if action_name == "High visibility":
            ws.cell(row=row, column=2).value = f'=IF(CONFIG!C27="YES","High visibility","")'
        else:
            ws.cell(row=row, column=2).value = action_name
        apply_style(ws.cell(row=row, column=2), font=Font(name="Aptos", size=11, bold=True),
                    border=BORDER_THIN)

        # Column C: Destructive
        if action_name == "High visibility":
            ws.cell(row=row, column=3).value = f'=IF(CONFIG!C27="YES","YES","")'
        else:
            ws.cell(row=row, column=3).value = destructive
        apply_style(ws.cell(row=row, column=3), font=FONT_NORMAL, alignment=ALIGN_CENTER,
                    border=BORDER_THIN)

        # Column D: Samples used (text description)
        if action_name == "Size set control":
            # Uses up to 3 samples from different sizes: min, base, max
            ws.cell(row=row, column=4).value = (
                '=IF(CONFIG!$F$18=0,"",'
                'LET(samples,FILTER(SAMPLES!$B$2:$C$31,SAMPLES!$A$2:$A$31<>""),'
                'colours,INDEX(samples,,1),sizes,INDEX(samples,,2),'
                'minS,CONFIG!$F$38,baseS,CONFIG!$F$40,maxS,CONFIG!$F$39,'
                'r1,IFERROR(INDEX(colours,MATCH(minS,sizes,0))&" / "&minS,""),'
                'r2,IF(baseS<>minS,IFERROR(", "&INDEX(colours,MATCH(baseS,sizes,0))&" / "&baseS,""),""),'
                'r3,IF(AND(maxS<>minS,maxS<>baseS),IFERROR(", "&INDEX(colours,MATCH(maxS,sizes,0))&" / "&maxS,""),""),'
                'r1&r2&r3))'
            )
        elif action_name == "Wash tests":
            ws.cell(row=row, column=4).value = (
                '=IF(CONFIG!$F$18=0,"",'
                'TEXTJOIN(", ",TRUE,'
                'IF(SAMPLES!$D$2:$D$31="Wash tests",'
                'SAMPLES!$B$2:$B$31&" / "&SAMPLES!$C$2:$C$31,"")))'
            )
        elif action_name == "Personalization guide":
            ws.cell(row=row, column=4).value = (
                '=IF(CONFIG!$F$18=0,"",'
                'TEXTJOIN(", ",TRUE,'
                'IF(SAMPLES!$D$2:$D$31="Personalization guide",'
                'SAMPLES!$B$2:$B$31&" / "&SAMPLES!$C$2:$C$31,"")))'
            )
        elif action_name == "High visibility":
            ws.cell(row=row, column=4).value = (
                '=IF(CONFIG!C27<>"YES","",'
                'TEXTJOIN(", ",TRUE,'
                'IF(SAMPLES!$D$2:$D$31="High visibility",'
                'SAMPLES!$B$2:$B$31&" / "&SAMPLES!$C$2:$C$31,"")))'
            )
        elif action_name == "Photo shooting":
            ws.cell(row=row, column=4).value = (
                '=IF(CONFIG!$F$18=0,"",'
                'TEXTJOIN(", ",TRUE,'
                'IF(SAMPLES!$D$2:$D$31="Photo shooting",'
                'SAMPLES!$B$2:$B$31&" / "&SAMPLES!$C$2:$C$31,"")))'
            )

        apply_style(ws.cell(row=row, column=4), font=FONT_NORMAL, alignment=ALIGN_WRAP,
                    border=BORDER_THIN)

        # Column E: Remaining stock
        # Stock = Total - cumulative destructive samples consumed so far
        if action_name == "Size set control":
            # Non-destructive: stock = total
            ws.cell(row=row, column=5).value = '=IF(CONFIG!$F$18=0,"",CONFIG!$F$18)'
        elif action_name == "Wash tests":
            # Destructive: stock = total - wash test count
            ws.cell(row=row, column=5).value = (
                '=IF(CONFIG!$F$18=0,"",CONFIG!$F$18-COUNTIF(SAMPLES!$D$2:$D$31,"Wash tests"))'
            )
        elif action_name == "Personalization guide":
            # Destructive: stock = total - wash - perso
            ws.cell(row=row, column=5).value = (
                '=IF(CONFIG!$F$18=0,"",CONFIG!$F$18'
                '-COUNTIF(SAMPLES!$D$2:$D$31,"Wash tests")'
                '-COUNTIF(SAMPLES!$D$2:$D$31,"Personalization guide"))'
            )
        elif action_name == "High visibility":
            ws.cell(row=row, column=5).value = (
                '=IF(CONFIG!C27<>"YES","",CONFIG!$F$18'
                '-COUNTIF(SAMPLES!$D$2:$D$31,"Wash tests")'
                '-COUNTIF(SAMPLES!$D$2:$D$31,"Personalization guide")'
                '-COUNTIF(SAMPLES!$D$2:$D$31,"High visibility"))'
            )
        elif action_name == "Photo shooting":
            # Non-destructive: same stock as previous step
            ws.cell(row=row, column=5).value = (
                '=IF(CONFIG!$F$18=0,"",CONFIG!$F$18'
                '-COUNTIF(SAMPLES!$D$2:$D$31,"Wash tests")'
                '-COUNTIF(SAMPLES!$D$2:$D$31,"Personalization guide")'
                '-IF(CONFIG!C27="YES",COUNTIF(SAMPLES!$D$2:$D$31,"High visibility"),0))'
            )

        apply_style(ws.cell(row=row, column=5), font=Font(name="Aptos", size=11, bold=True),
                    alignment=ALIGN_CENTER, border=BORDER_THIN)

    # Conditional formatting: destructive cells
    ws.conditional_formatting.add(
        "C5:C9",
        CellIsRule(operator="equal", formula=['"YES"'],
                   fill=FILL_DESTRUCTIVE, font=Font(color="C62828", bold=True))
    )
    ws.conditional_formatting.add(
        "C5:C9",
        CellIsRule(operator="equal", formula=['"NO"'],
                   fill=FILL_SAFE, font=Font(color="2E7D32", bold=True))
    )

    # Hide row if empty (high visibility row when not active)
    ws.conditional_formatting.add(
        "A8:E8",
        FormulaRule(
            formula=['$B8=""'],
            font=Font(color="FFFFFF"),
            fill=FILL_WHITE,
            border=Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="thin", color="FFFFFF"),
            )
        )
    )

    # Stock bar: conditional formatting for remaining stock (green gradient)
    ws.conditional_formatting.add(
        "E5:E9",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        "E5:E9",
        CellIsRule(operator="equal", formula=["0"],
                   fill=PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
                   font=Font(color="C62828", bold=True))
    )

    return ws


# ============================================================
# MAIN: Generate the workbook
# ============================================================

def main():
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create sheets in order
    create_ref_sheet(wb)
    ws_config = create_config_sheet(wb)
    create_samples_sheet(wb)
    create_order_grid_sheet(wb)
    create_sample_flow_sheet(wb)

    # Set CONFIG as the active sheet (first visible)
    wb.active = wb.sheetnames.index("CONFIG")

    # Hide REF sheet (optional - user can unhide if curious)
    wb["REF"].sheet_state = "hidden"

    # Save
    output_path = r"C:\Users\Dell\Documents\01_developpement\05_Samples_management\Samples_Management_Tool.xlsx"
    wb.save(output_path)
    print(f"Excel file generated: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print("Open the file in Excel 365 and fill in the CONFIG sheet.")


if __name__ == "__main__":
    main()
